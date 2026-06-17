import openpyxl

def main():
    wb = openpyxl.load_workbook('calculation.xlsx', data_only=False)
    target_sheet = 'STORAGE -  SHUTTER '
        
    print(f"Reading sheet: {target_sheet}")
    ws = wb[target_sheet]
    
    for row in range(24, 46):
        cell_e = ws[f'E{row}']
        cell_f = ws[f'F{row}']
        cell_g = ws[f'G{row}']
        
        print(f"Row {row} | Qty: {cell_e.value} | Rate: {cell_f.value} | Amount: {cell_g.value}")

if __name__ == '__main__':
    main()
